from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Notice, Event, News, ListPart, Participant, Rating, \
    Organizer, Purpose, Criteria
from django.db import models
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User, Group
from django.forms import ModelForm, HiddenInput, TextInput, formset_factory
from django.db.models import Max, Min, Avg
from django import forms
from datetime import date
from datetime import datetime
from bootstrap_datepicker_plus import DatePickerInput, TimePickerInput
from openpyxl import Workbook
from django.http import HttpResponse


def group_is(user, list_of_groups):
    return user.groups.filter(name__in=list_of_groups).exists()


def index(request):
    today = date.today()
    posts = Notice.objects.all().select_related('id_event').filter(date_notice__gte=today)
    purposes = Purpose.objects.all()

    args = {
        'posts': posts,
        'purposes': purposes,
    }
    return render(request, 'index.html', args)


def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)

        if form.is_valid():
            user = form.save()

            if request.POST['is_org'] == 'on':
                group = Group.objects.get(name='Организации')
            else:
                group = Group.objects.get(name='Участники')

            user.groups.add(group)

            return redirect('index')
        else:
            print(form.errors)
            return render(request, '404.html')
    else:
        form = UserCreationForm()

        args = {'form': form}
        return render(request, 'auth/register.html', args)


class UserLogin(LoginView):
    template_name = 'auth/login.html'


class UserLogout(LoginRequiredMixin, LogoutView):
    template_name = 'auth/logout.html'


@login_required
def dashboard(request):
    args = {'content': ''}

    if group_is(request.user, ['Администратор']):
        args['categories'] = [
            {
                'name': 'Отчет о событиях',
                'icon_class': 'bx-calendar',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'admin_report',
            },
        ]
    elif group_is(request.user, ['Организации']):
        args['categories'] = [
            {
                'name': 'События',
                'icon_class': 'bxs-calendar-event',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'events_list',
            },
            {
                'name': 'Мероприятия',
                'icon_class': 'bxs-spreadsheet',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'notices_list',
            },
            {
                'name': 'Списки участников',
                'icon_class': 'bxs-user-badge',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'part_list',
            },
            {
                'name': 'Новости',
                'icon_class': 'bxs-news',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'news_list',
            },
            {
                'name': 'Оценки',
                'icon_class': 'bxs-star-half',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'evaluation',
            },
            {
                'name': 'Личная информация',
                'icon_class': 'bxs-user',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'user_info',
            },
        ]

    elif group_is(request.user, ['Участники']):
        args['categories'] = [
            {
                'name': 'Мои события',
                'icon_class': 'bxs-calendar-event',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'parts_events',
            },
            {
                'name': 'Мои мероприятия',
                'icon_class': 'bxs-spreadsheet',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'parts_notices',
            },
            {
                'name': 'Мои оценки',
                'icon_class': 'bxs-star-half',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'parts_evaluation',
            },
            {
                'name': 'Личная информация',
                'icon_class': 'bxs-user',
                'description': 'Lorem ipsum dolor sit amet consectetur adipisicing elit.',
                'url': 'user_info',
            },
        ]

    return render(request, 'dashboard/dashboard.html', args)


class OrganizerForm(ModelForm):
    class Meta:
        model = Organizer
        fields = [
            'full_name_org',
            'short_name_org',
            'address',
            'full_name_head',
            'email_org',
            'phone_org',
            'inn',
            'ogrn_ogrnip',
        ]
        labels = {
            'full_name_org': 'Полное название организации',
            'short_name_org': 'Короткое название',
            'address': 'Адрес',
            'full_name_head': 'ФИО руководителя',
            'email_org': 'Email',
            'phone_org': 'Телефон',
            'inn': 'ИНН',
            'ogrn_ogrnip': 'ОГРН/ОГРНИП',
        }
        widgets = {
            'login': HiddenInput(),
        }


class ParticipantForm(ModelForm):
    class Meta:
        model = Participant
        fields = [
            'full_name_part',
            'date_birth',
            'email_part',
            'passport',
            'phone_part',
        ]
        labels = {
            'full_name_part': 'ФИО',
            'date_birth': 'Дата рождения',
            'email_part': 'Email',
            'passport': 'Серия и номер паспорта',
            'phone_part': 'Телефон',
        }
        widgets = {
            'login': HiddenInput(),
        }

@login_required
def user_info(request):
    if group_is(request.user, ['Участники']):
        user = Participant.objects.get(login=request.user.id)

        if request.method == 'POST':
            form = ParticipantForm(request.POST, instance=user)

            if form.is_valid():
                form.save()

                return redirect('dashboard')

            return render(request, '404.html')

        else:
            form = ParticipantForm(instance=user)

            args = {
                'form': form,
                'title': 'Форма изменения личных данных',
                'desc': 'Здесь вы сможете изменять данные своего пользователя',
            }

            return render(request, 'dashboard/general_form.html', args)

    elif group_is(request.user, ['Организации']):
        user = Organizer.objects.get(login=request.user.id)

        if request.method == 'POST':
            form = OrganizerForm(request.POST, instance=user)

            if form.is_valid():
                form.save()

                return redirect('dashboard')

            return render(request, '404.html')

        else:
            form = OrganizerForm(instance=user)

            args = {
                'form': form,
                'title': 'Форма изменения личных данных',
                'desc': 'Здесь вы сможете изменять данные своей организации',
            }

            return render(request, 'dashboard/general_form.html', args)


class ReportForm(forms.Form):
    begin_date = forms.DateField(
        widget=DatePickerInput(
            options={
                "format": "DD.MM.YYYY",
                "locale": "ru",
            }
        ),
        label='Начальная дата',
        required=False
    )
    end_date = forms.DateField(
        widget=DatePickerInput(
            options={
                "format": "DD.MM.YYYY",
                "locale": "ru",
            }
        ),
        label='Конечная дата',
        required=False
    )

@login_required
def admin_report(request):
    events_evals = None
    begin_date = None
    end_date = None

    if not(group_is(request.user, ['Администратор'])):
        return render(request, '404.html')

    if request.method == 'POST':
        if request.POST['begin_date']:
            begin_date = request.POST['begin_date']
            begin_date = begin_date[6:] + '-' + begin_date[3:5] + '-' + begin_date[:2]
        if request.POST['end_date']:
            end_date = request.POST['end_date']
            end_date = end_date[6:] + '-' + end_date[3:5] + '-' + end_date[:2]

        form = ReportForm(request.POST)

        if form.is_valid():
            if begin_date and not(end_date)
                events_evals = get_report(begin_date, None)
            elif not(begin_date) and end_date:
                events_evals = get_report(None, end_date)
            elif begin_date and end_date:
                events_evals = get_report(begin_date, end_date)
            else:
                events_evals = get_report(None, None)
        else:
            print(form.errors, begin_date, end_date)
            return render(request, '404.html')
    else:
        form = ReportForm()

        events_evals = get_report(None, None)

    if request.method == 'POST' and request.POST['action'] == 'XLS':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = 'Отчет'

        columns = [
            'Мероприятие',
            'Событие',
            'Количество участников',
            'Дата',
            'Оценка',
        ]

        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for item in events_evals:
            for notice in item['notices_evals']:
                row = [
                    str(notice['notice_info']),
                    str(item['event']),
                    str(notice['people_counter']),
                    str(notice['notice_date']),
                    str(notice['efficiency']),
                ]
                row_num += 1

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    args = {
        'name': 'Оценки',
        'form': form,
        'events_evals': events_evals,
    }
    return render(request, 'dashboard/admin_report.html', args)


def get_report(begin_date, end_date):
    exist_notices = Rating.objects.all().values_list('id_notice', flat=True)
    exist_events = Notice.objects.filter(id_notice__in=exist_notices).values_list('id_event', flat=True)

    events = Event.objects.filter(id_event__in=exist_events)
    criteria_list = Criteria.objects.all()

    events_evals = []

    for event in events:
        if begin_date and end_date is None:
            notices = Notice.objects.filter(id_event=event, id_notice__in=exist_notices, date_notice__gte=begin_date)
        elif begin_date is None and end_date:
            notices = Notice.objects.filter(id_event=event, id_notice__in=exist_notices, date_notice__lte=end_date)
        elif begin_date and end_date:
            notices = Notice.objects.filter(id_event=event, id_notice__in=exist_notices, date_notice__gte=begin_date, date_notice__lte=end_date)
        else:
            notices = Notice.objects.filter(id_event=event, id_notice__in=exist_notices)

        notices_evals = []

        for notice in notices:
            efficiency = 0

            for criteria in criteria_list:
                rating = Rating.objects.filter(id_notice=notice, id_criteria=criteria)

                if not(rating):
                    break

                avg_rating = float(rating.aggregate(Avg('evaluation'))['evaluation__avg'])
                max_rating = float(criteria.maximum)
                min_rating = float(criteria.minimum)

                normalized_rating = (avg_rating - min_rating) / (max_rating - min_rating)

                efficiency += normalized_rating * float(criteria.weight)

            people_counter = ListPart.objects.filter(id_notice=notice).count()

            notices_evals.append({
                'notice_info': notice,
                'notice_date': notice.date_notice,
                'efficiency': efficiency,
                'people_counter': people_counter,
            })

        events_evals.append({
            'event': event.name_event,
            'notices_evals': notices_evals,
        })

    return events_evals


@login_required
def events(request):
    events = Event.objects.all()

    args = {
        'name': 'События',
        'description': 'Описание странички',
        'url': 'event_view',
        'is_part': group_is(request.user, ['Участники']),
        'records': events,
    }
    return render(request, 'lists.html', args)


def notices(request):
    notices = Notice.objects.all()

    args = {
        'name': 'Мероприятия',
        'description': 'Описание странички',
        'url': 'notice_view',
        'is_part': group_is(request.user, ['Участники']),
        'records': notices,
    }
    return render(request, 'lists.html', args)


def news(request):
    news = News.objects.all()

    args = {
        'name': 'Новости',
        'description': 'Описание странички',
        'url': 'news_view',
        'is_part': group_is(request.user, ['Участники']),
        'records': news,
    }
    return render(request, 'lists.html', args)


def evaluation(request):
    exist_notices = Rating.objects.all().values_list('id_notice', flat=True)
    exist_events = Notice.objects.filter(id_notice__in=exist_notices).values_list('id_event', flat=True)

    events = Event.objects.filter(login=request.user.id, id_event__in=exist_events)
    criteria_list = Criteria.objects.all()

    events_evals = []

    for event in events:

        notices = Notice.objects.filter(id_event=event, id_notice__in=exist_notices)

        notices_evals = []

        for notice in notices:
            bad_news = []
            efficiency = 0

            for criteria in criteria_list:
                rating = Rating.objects.filter(id_notice=notice, id_criteria=criteria)

                if not(rating):
                    break

                avg_rating = float(rating.aggregate(Avg('evaluation'))['evaluation__avg'])
                max_rating = float(criteria.maximum)
                min_rating = float(criteria.minimum)

                normalized_rating = (avg_rating - min_rating) / (max_rating - min_rating)

                efficiency += normalized_rating * float(criteria.weight)

                if normalized_rating < 0.7:
                    bad_news.append(criteria)

            notices_evals.append({
                'notice_info': notice,
                'efficiency': efficiency,
                'bad_news': bad_news,
            })

        events_evals.append({
            'event': event.name_event,
            'notices_evals': notices_evals,
        })

    args = {
        'name': 'Оценки',
        'events_evals': events_evals,
    }

    return render(request, 'dashboard/evalution.html', args)


def parts_events(request):
    notices_listpart = ListPart.objects.values_list('id_notice', flat=True).filter(login=request.user.id)
    events_notices = Notice.objects.filter(id_notice__in=notices_listpart).values_list('id_event', flat=True)
    events = Event.objects.filter(id_event__in=events_notices)

    args = {'events': events}
    return render(request, 'dashboard/parts/events.html', args)


def parts_notices(request):
    notices_listpart = ListPart.objects.values_list('id_notice', flat=True).filter(login=request.user.id)
    notices = Notice.objects.filter(id_notice__in=notices_listpart)

    args = {'notices': notices}
    return render(request, 'dashboard/parts/notices.html', args)


def parts_evaluation(request):
    notices_listpart = ListPart.objects.values_list('id_notice', flat=True).filter(login=request.user.id)
    notices = Notice.objects.filter(id_notice__in=notices_listpart)

    args = {'notices': notices}
    return render(request, 'dashboard/parts/notices_evals.html', args)


class RatingForm(ModelForm):
    class Meta:
        model = Rating
        fields = [
            'login',
            'id_notice',
            'id_criteria',
            'evaluation',
        ]
        labels = {
            'evaluation': '',
        }
        widgets = {
            'login': HiddenInput(),
            'id_notice': HiddenInput(),
            'id_criteria': HiddenInput(),
        }


def parts_eval_notice(request, notice_id):
    criteria_list = Criteria.objects.all()

    RatingFormSet = formset_factory(RatingForm, extra=0)

    if request.method == 'POST':
        formset = RatingFormSet(request.POST)

        if formset.is_valid():
            for form in formset:
                if form.is_valid():
                    form.save()
                else:
                    print(form.errors)

            return redirect('parts_evaluation')

        return render(request, '404.html')

    else:
        notice = Notice.objects.get(id_notice=notice_id)

        initial_data = []
        for criteria in criteria_list:
            initial_data.append({
                'login': request.user.id,
                'id_notice': notice,
                'id_criteria': criteria,
                'evaluation': 1
            })

        formset = RatingFormSet(initial=initial_data)

        args = {
            'formset': formset,
            'title': 'Форма изменения записи',
            'desc': 'Здесь вы сможете изменять детали конкретного События',
            'criteria_list': criteria_list,
            'zip_list': zip(formset, criteria_list),
        }

        return render(request, 'dashboard/general_formset.html', args)


def event_view(request, record_id):
    event = Event.objects.get(id_event=record_id)

    args = {
        'event': event,
        'is_part': group_is(request.user, ['Участники']),
    }
    return render(request, 'record/event.html', args)


def notice_view(request, record_id):
    notice = Notice.objects.get(id_notice=record_id)

    remains = notice.max_part - ListPart.objects.filter(id_notice=notice).count()

    is_part = group_is(request.user, ['Участники'])
    is_part_of_current = True if ListPart.objects.filter(
        id_notice=notice, login=request.user.id) else False

    args = {
        'notice': notice,
        'is_part': is_part,
        'is_part_of_current': is_part_of_current,
        'remains': remains,
    }
    return render(request, 'record/notice.html', args)


def news_view(request, record_id):
    news = News.objects.get(id_news=record_id)

    args = {
        'news': news,
        'is_part': group_is(request.user, ['Участники']),
    }
    return render(request, 'record/news.html', args)


def become_part(request, notice_id):
    user_id = request.user.id

    if ListPart.objects.filter(id_notice=notice_id, login=user_id):
        ListPart.objects.filter(id_notice=notice_id,
                                login=user_id).delete()
    else:
        part = Participant.objects.get(login=user_id)
        notice = Notice.objects.get(id_notice=notice_id)

        ListPart.objects.create(login=part,
            id_notice=notice,
            date_record=str(date.today()),
        )

    return redirect('notice_view', record_id=notice_id)


def unfollow_notice(request, notice_id):
    user_id = request.user.id

    ListPart.objects.filter(id_notice=notice_id, login=user_id).delete()

    return redirect('parts_notices')


@login_required
def events_list(request):
    events = Event.objects.filter(login=request.user.id)

    args = {'events': events}
    return render(request, 'dashboard/events.html', args)


class EventForm(ModelForm):
    class Meta:
        model = Event
        fields = [
            'id_event',
            'login',
            'name_event',
            'document',
            'logo',
            'id_level',
            'id_type',
        ]
        labels = {
            'name_event': 'Название',
            'id_level': 'Уровень',
            'id_type': 'Тип',
            'document': 'Документ',
            'logo': 'Лого',
        }
        widgets = {
            'id_event': HiddenInput(),
            'login': HiddenInput(),
        }


@login_required
def event_edit(request, event_id):
    event = Event.objects.get(id_event=event_id)

    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)

        if form.is_valid():
            form.save()

            return redirect('events_list')

        return redirect('event_edit', event_id=event_id)

    else:
        form = EventForm(instance=event)

        args = {
            'form': form,
            'title': 'Форма изменения записи',
            'desc': 'Здесь вы сможете изменять детали конкретного События',
        }

        return render(request, 'dashboard/general_form.html', args)


@login_required
def event_add(request):
    init_login = request.user.id
    init_id_event = list(Event.objects.all().aggregate(
        Max('id_event')).values())[0] + 1

    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)

        if form.is_valid() and form['login'].value() == str(init_login) and form['id_event'].value() == str(init_id_event):
            form.save()

            return redirect('events_list')
        else:
            return redirect('event_add')

    else:
        initial = {
            'id_event': init_id_event,
            'login': init_login,
            'document': '',
            'logo': '',
        }

        form = EventForm(initial)

        args = {
            'form': form,
            'title': 'Форма добавления записи',
            'desc': 'Здесь вы можете добавлять События',
        }

        return render(request, 'dashboard/general_form.html', args)


@login_required
def event_remove(request, event_id):
    Event.objects.get(id_event=event_id).delete()

    return redirect('events_list')


@login_required
def notices_list(request):
    events = Event.objects.filter(login=request.user.id)
    notices = Notice.objects.filter(id_event__in=events)

    args = {'notices': notices}
    return render(request, 'dashboard/notices.html', args)


class NoticeForm(ModelForm):
    class Meta:
        model = Notice
        fields = [
            'id_notice',
            'id_event',
            'id_purpose',
            'date_notice',
            'max_part',
            'description',
            'time_n',
            'location',
            'status',
        ]
        labels = {
            'id_event': 'Событие',
            'id_purpose': 'Цель',
            'date_notice': 'Дата',
            'max_part': 'Макс. кол-во участников',
            'description': 'Описание',
            'time_n': 'Время начала',
            'location': 'Место проведения',
            'status': 'Статус',
        }
        widgets = {
            'id_notice': HiddenInput(),
            'date_notice': DatePickerInput(
                options={
                    "format": "DD.MM.YYYY",
                    "locale": "ru",
                }
            ),
            'time_n': TimePickerInput(),
        }


@login_required
def notice_edit(request, notice_id):
    notice = Notice.objects.get(id_notice=notice_id)
    init_id_event = Event.objects.filter(
            login=request.user.id)

    if request.method == 'POST':
        form = NoticeForm(request.POST, instance=notice)
        form.fields['id_event'].queryset = init_id_event

        if form.is_valid():
            form.save()

            return redirect('notices_list')

        return redirect('notice_edit', notice_id=notice_id)

    else:
        form = NoticeForm(instance=notice)
        form.fields['id_event'].queryset = init_id_event

        args = {
            'form': form,
            'title': 'Форма изменения записи',
            'desc': 'Здесь вы сможете изменять детали конкретного Мероприятия',
        }

        return render(request, 'dashboard/general_form.html', args)


@login_required
def notice_add(request):
    init_id_notice = list(Notice.objects.all().aggregate(
        Max('id_notice')).values())[0] + 1
    init_id_event = Event.objects.filter(
        login=request.user.id)

    if request.method == 'POST':
        form = NoticeForm(request.POST)
        form.fields['id_event'].queryset = init_id_event

        if form.is_valid() and form['id_notice'].value() == str(init_id_notice):
            form.save()

            return redirect('notices_list')
        else:
            return redirect('notice_add')

    else:
        initial = {
            'id_notice': init_id_notice,
        }

        form = NoticeForm(initial)
        form.fields['id_event'].queryset = init_id_event

        args = {
            'form': form,
            'title': 'Форма добавления записи',
            'desc': 'Здесь вы можете добавлять Мероприятия',
        }

        return render(request, 'dashboard/general_form.html', args)


@login_required
def notice_remove(request, notice_id):
    Notice.objects.get(id_notice=notice_id).delete()

    return redirect('notices_list')


@login_required
def news_list(request):
    events = Event.objects.filter(login=request.user.id)
    notices = Notice.objects.filter(id_event__in=events)
    news = News.objects.filter(id_notice__in=notices)

    args = {'news': news}
    return render(request, 'dashboard/news.html', args)


class NewsForm(ModelForm):
    class Meta:
        model = News
        fields = [
            'id_news',
            'information',
            'photos',
            'id_notice',
            'date_news',
        ]
        labels = {
            'information': 'Информация',
            'photos': 'Файл изображения',
            'id_notice': 'Мероприятие',
        }
        widgets = {
            'id_news': HiddenInput(),
            'date_news': HiddenInput(),
        }


@login_required
def news_edit(request, news_id):
    news = News.objects.get(id_news=news_id)

    events = Event.objects.filter(login=request.user.id)
    init_id_notice = Notice.objects.filter(id_event__in=events)

    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES, instance=news)
        form.fields['id_notice'].queryset = init_id_notice

        print(form.is_valid(), form.errors)

        if form.is_valid():
            form.save()

            return redirect('news_list')

        return redirect('news_edit', news_id=news_id)
    else:
        form = NewsForm(instance=news)

        args = {
            'form': form,
            'title': 'Форма изменения записи',
            'desc': 'Здесь вы сможете изменять детали конкретной Новости',
        }
        return render(request, 'dashboard/general_form.html', args)


@login_required
def news_add(request):
    init_id_news = list(News.objects.all().aggregate(
        Max('id_news')).values())[0] + 1
    init_date_news = str(date.today())

    events = Event.objects.filter(login=request.user.id)
    init_id_notices = Notice.objects.filter(id_event__in=events)

    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES)
        form.fields['id_notice'].queryset = init_id_notices

        if form.is_valid() and form['id_news'].value() == str(init_id_news) \
                and form['date_news'].value() == init_date_news:
            form.save()

            return redirect('news_list')
        else:
            return redirect('news_add')
    else:
        initial = {
            'id_news': init_id_news,
            'date_news': init_date_news,
        }

        form = NewsForm(initial)
        form.fields['id_notice'].queryset = init_id_notices

        args = {
            'form': form,
            'title': 'Форма добавления записи',
            'desc': 'Здесь вы можете добавлять Новости',
        }

        return render(request, 'dashboard/general_form.html', args)


@login_required
def news_remove(request, news_id):
    News.objects.get(id_news=news_id).delete()

    return redirect('news_list')


@login_required
def part_list(request):
    events = Event.objects.filter(login=request.user.id)
    notices = Notice.objects.filter(id_event__in=events)
    lists = ListPart.objects.filter(id_notice__in=notices)

    args = {'lists': lists}
    return render(request, 'dashboard/part_list.html', args)
